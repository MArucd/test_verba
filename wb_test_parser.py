"""Simple Wildberries parser for a test assignment.

Creates:
1) Full catalog XLSX for a search query
2) Filtered XLSX (rating >= 4.5, price <= 10000, country of production = Russia)
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import time
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment

SEARCH_WB_URL = "https://search.wb.ru/exactmatch/ru/common/v18/search"
SEARCH_GOODS_URL = "https://search-goods.wildberries.ru/search"
CARD_V4_URL = "https://card.wb.ru/cards/v4/detail"
WB_HOME_URL = "https://www.wildberries.ru/"
CHROME_FOR_TESTING_JSON_URL = (
    "https://googlechromelabs.github.io/"
    "chrome-for-testing/known-good-versions-with-downloads.json"
)
TOKEN_COOKIE_NAME = "x_wbaas_token"
WB_TOKEN_DOMAINS = (
    ".wildberries.ru",
    ".wb.ru",
    ".wbbasket.ru",
    ".search.wb.ru",
    ".card.wb.ru",
)
CHROMEDRIVER_CACHE_DIR = Path.home() / ".cache" / "wb_test_parser" / "chromedriver"

DEFAULT_QUERY = "пальто из натуральной шерсти"
DEFAULT_OUTPUT_ALL = "wb_catalog.xlsx"
DEFAULT_OUTPUT_FILTERED = "wb_catalog_filtered.xlsx"

REQUEST_HEADERS = {
    "accept": "*/*",
    "accept-language": "ru-RU,ru;q=0.9",
    "user-agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/142.0.0.0 Safari/537.36"
    ),
    "x-requested-with": "XMLHttpRequest",
}

VOL_HOST_BREAKPOINTS: list[tuple[int, str]] = [
    (143, "01"),
    (287, "02"),
    (431, "03"),
    (719, "04"),
    (1007, "05"),
    (1061, "06"),
    (1115, "07"),
    (1169, "08"),
    (1313, "09"),
    (1601, "10"),
    (1655, "11"),
    (1919, "12"),
    (2045, "13"),
    (2189, "14"),
    (2405, "15"),
    (2621, "16"),
    (2837, "17"),
    (3053, "18"),
    (3269, "19"),
    (3485, "20"),
    (3701, "21"),
    (3917, "22"),
    (4133, "23"),
    (4349, "24"),
    (4565, "25"),
    (4877, "26"),
    (5189, "27"),
    (5501, "28"),
    (5813, "29"),
    (6125, "30"),
    (6437, "31"),
    (6749, "32"),
    (7061, "33"),
    (7373, "34"),
    (7685, "35"),
    (7997, "36"),
    (8309, "37"),
    (8741, "38"),
    (9173, "39"),
    (9605, "40"),
]

COLUMNS = [
    "Ссылка на товар",
    "Артикул",
    "Название",
    "Цена",
    "Описание",
    "Ссылки на изображения через запятую",
    "Все характеристики с сохранением структуры",
    "Название селлера",
    "Ссылка на селлера",
    "Размеры товара через запятую",
    "Остатки по товару (число)",
    "Рейтинг",
    "Количество отзывов",
    "Страна производства",
]


@dataclass
class StaticBundle:
    """Static files loaded from `basket-XX.wbbasket.ru`."""

    base_url: str
    card: dict[str, Any]
    seller: dict[str, Any]
    price_history: list[dict[str, Any]]


def json_from_response(response: requests.Response) -> Any | None:
    """Parse JSON even if WB returns `text/plain` content type."""
    content_type = response.headers.get("content-type", "").lower()
    body = response.text.strip()

    if not body:
        return None

    maybe_json = "json" in content_type or body[0] in "{["
    if not maybe_json:
        return None

    try:
        return response.json()
    except ValueError:
        return None


def request_json(
    session: requests.Session,
    url: str,
    *,
    params: dict[str, Any] | None = None,
    method: str = "GET",
    json_body: Any | None = None,
    headers: dict[str, str] | None = None,
    retries: int = 3,
    timeout: int = 25,
) -> Any | None:
    # pylint: disable=too-many-arguments
    """Request JSON with retries and lightweight backoff."""
    req_headers = REQUEST_HEADERS.copy()
    if headers:
        req_headers.update(headers)

    for attempt in range(1, retries + 1):
        try:
            if method == "POST":
                response = session.post(
                    url,
                    params=params,
                    json=json_body,
                    headers=req_headers,
                    timeout=timeout,
                )
            else:
                response = session.get(
                    url,
                    params=params,
                    headers=req_headers,
                    timeout=timeout,
                )
        except requests.RequestException:
            time.sleep(0.8 * attempt)
            continue

        status_code = response.status_code
        if status_code in {400, 401, 403, 404}:
            return None

        if response.status_code == 200:
            parsed = json_from_response(response)
            if parsed is not None:
                return parsed

        time.sleep(0.8 * attempt)

    return None


def parse_version(version: str) -> tuple[int, int, int, int]:
    """Convert `x.y.z.w` into tuple for sorting."""
    numbers = [int(part) for part in re.findall(r"\d+", version)[:4]]
    while len(numbers) < 4:
        numbers.append(0)
    return numbers[0], numbers[1], numbers[2], numbers[3]


def detect_local_chrome_version() -> str | None:
    """Detect installed Google Chrome version."""
    commands = (
        ["google-chrome", "--version"],
        ["chrome", "--version"],
        ["chromium-browser", "--version"],
        ["chromium", "--version"],
    )

    for command in commands:
        process = subprocess.run(
            command,
            capture_output=True,
            text=True,
            check=False,
        )
        if process.returncode != 0:
            continue

        match = re.search(
            r"(\d+\.\d+\.\d+\.\d+)",
            f"{process.stdout} {process.stderr}",
        )
        if match:
            return match.group(1)

    return None


def select_chromedriver_linux_url(  # pylint: disable=too-many-branches
    chrome_version: str,
) -> tuple[str, str]:
    """Select closest chromedriver URL for local Chrome build."""
    response = requests.get(CHROME_FOR_TESTING_JSON_URL, timeout=30)
    response.raise_for_status()
    data = response.json()
    versions = data.get("versions")
    if not isinstance(versions, list):
        raise RuntimeError("Не удалось получить список версий chromedriver")

    parts = chrome_version.split(".")
    major_prefix = parts[0]
    build_prefix = ".".join(parts[:3])

    candidates: list[dict[str, Any]] = []
    for item in versions:
        if not isinstance(item, dict):
            continue
        version = str(item.get("version", ""))
        if version.startswith(f"{build_prefix}."):
            candidates.append(item)

    if not candidates:
        for item in versions:
            if not isinstance(item, dict):
                continue
            version = str(item.get("version", ""))
            if version.startswith(f"{major_prefix}."):
                candidates.append(item)

    if not candidates:
        raise RuntimeError("Не удалось подобрать chromedriver по версии Chrome")

    selected = max(
        candidates,
        key=lambda item: parse_version(str(item.get("version", "0.0.0.0"))),
    )
    selected_version = str(selected.get("version", ""))
    downloads = (selected.get("downloads") or {}).get("chromedriver", [])
    if not isinstance(downloads, list):
        raise RuntimeError("Нет секции downloads/chromedriver для выбранной версии")

    for download in downloads:
        if not isinstance(download, dict):
            continue
        if download.get("platform") == "linux64" and isinstance(
            download.get("url"),
            str,
        ):
            return selected_version, str(download["url"])

    raise RuntimeError("Не найден linux64 chromedriver в списке загрузок")


def ensure_chromedriver(chromedriver_path: str | None = None) -> Path:
    """Return path to executable chromedriver (download if needed)."""
    if chromedriver_path:
        path = Path(chromedriver_path).expanduser().resolve()
        if not path.exists():
            raise RuntimeError(f"chromedriver не найден: {path}")
        return path

    chrome_version = detect_local_chrome_version()
    if not chrome_version:
        raise RuntimeError("Не удалось определить локальную версию Google Chrome")

    selected_version, driver_url = select_chromedriver_linux_url(chrome_version)
    target_dir = CHROMEDRIVER_CACHE_DIR / selected_version
    driver_path = target_dir / "chromedriver-linux64" / "chromedriver"

    if driver_path.exists():
        return driver_path

    target_dir.mkdir(parents=True, exist_ok=True)
    archive_path = target_dir / "chromedriver-linux64.zip"

    response = requests.get(driver_url, timeout=90)
    response.raise_for_status()
    archive_path.write_bytes(response.content)

    with zipfile.ZipFile(archive_path) as archive:
        archive.extractall(target_dir)

    if not driver_path.exists():
        raise RuntimeError("Chromedriver скачан, но исполняемый файл не найден")

    driver_path.chmod(0o755)
    return driver_path


def get_wb_token(  # pylint: disable=import-outside-toplevel
    chromedriver_path: Path,
    timeout_seconds: int = 30,
) -> str:
    """Get WB token (`x_wbaas_token`) from browser cookies."""
    # Local import so parser still imports even when Selenium is absent.
    from selenium import webdriver  # pylint: disable=import-outside-toplevel
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service

    options = Options()
    options.binary_location = "/usr/bin/google-chrome"
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1280,900")
    options.add_argument(f"--user-agent={REQUEST_HEADERS['user-agent']}")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    service = Service(executable_path=str(chromedriver_path))
    driver = webdriver.Chrome(service=service, options=options)
    try:
        driver.execute_cdp_cmd("Network.enable", {})
        driver.get(WB_HOME_URL)
        deadline = time.time() + timeout_seconds

        while time.time() < deadline:
            cookies_data = driver.execute_cdp_cmd("Network.getAllCookies", {})
            cookies = cookies_data.get("cookies") or []
            for cookie in cookies:
                if not isinstance(cookie, dict):
                    continue
                if cookie.get("name") == TOKEN_COOKIE_NAME:
                    token = cookie.get("value")
                    if isinstance(token, str) and token.strip():
                        return token
            time.sleep(1.0)
    finally:
        driver.quit()

    raise RuntimeError("Не удалось получить x_wbaas_token через браузер")


def apply_wb_token(session: requests.Session, token: str) -> None:
    """Apply token cookie to all WB domains we use."""
    for domain in WB_TOKEN_DOMAINS:
        session.cookies.set(TOKEN_COOKIE_NAME, token, domain=domain, path="/")


def basket_host_by_nm_id(nm_id: int) -> str:
    """Build `basket-XX.wbbasket.ru/volN` host according to WB mapping."""
    vol = nm_id // 100000
    shard = "41"
    for upper_bound, bucket in VOL_HOST_BREAKPOINTS:
        if vol <= upper_bound:
            shard = bucket
            break
    return f"basket-{shard}.wbbasket.ru/vol{vol}"


def static_base_url(nm_id: int) -> str:
    """Build base URL for static product files."""
    host = basket_host_by_nm_id(nm_id)
    part = nm_id // 1000
    return f"https://{host}/part{part}/{nm_id}"


def get_catalog_ids_from_search_wb(
    session: requests.Session,
    query: str,
    max_pages: int = 10,
) -> list[int]:
    """Try to collect IDs from `search.wb.ru` (best quality search source)."""
    ids: list[int] = []
    seen: set[int] = set()

    base_params = {
        "ab_testing": "false",
        "appType": "1",
        "curr": "rub",
        "dest": "-1257786",
        "hide_dtype": "13",
        "lang": "ru",
        "query": query,
        "resultset": "catalog",
        "sort": "popular",
        "spp": "30",
        "suppressSpellcheck": "false",
    }

    for page in range(1, max_pages + 1):
        params = base_params | {"page": str(page)}
        data = request_json(session, SEARCH_WB_URL, params=params, retries=4)
        if not isinstance(data, dict):
            if page == 1:
                return []
            break

        products = data.get("products") or []
        if not products:
            break

        for product in products:
            nm_id = product.get("id")
            if isinstance(nm_id, int) and nm_id not in seen:
                seen.add(nm_id)
                ids.append(nm_id)

        if len(products) < 100:
            break

        time.sleep(0.5)

    return ids


def get_catalog_ids_from_search_goods(
    session: requests.Session,
    query: str,
) -> list[int]:
    """Fallback ID source when `search.wb.ru` is throttled."""
    data = request_json(
        session,
        SEARCH_GOODS_URL,
        params={"query": query},
        retries=3,
    )
    if not isinstance(data, list):
        return []

    unique_ids: list[int] = []
    seen: set[int] = set()
    for item in data:
        if isinstance(item, int) and item not in seen:
            seen.add(item)
            unique_ids.append(item)
    return unique_ids


def get_catalog_ids(session: requests.Session, query: str) -> tuple[list[int], str]:
    """Fetch IDs from primary source, then fallback if needed."""
    primary = get_catalog_ids_from_search_wb(session, query=query)
    if primary:
        return primary, "search.wb.ru"

    fallback = get_catalog_ids_from_search_goods(session, query=query)
    return fallback, "search-goods.wildberries.ru"


def get_card_v4(session: requests.Session, nm_id: int) -> dict[str, Any] | None:
    """Load card data with stocks, rating and reviews."""
    params = {
        "appType": "1",
        "curr": "rub",
        "dest": "-1257786",
        "spp": "30",
        "nm": str(nm_id),
    }
    data = request_json(session, CARD_V4_URL, params=params, retries=3)
    if not isinstance(data, dict):
        return None

    products = data.get("products")
    if not isinstance(products, list) or not products:
        return None

    first = products[0]
    if not isinstance(first, dict):
        return None
    return first


def get_static_bundle(session: requests.Session, nm_id: int) -> StaticBundle | None:
    """Load static card/seller/price history files from basket host."""
    base = static_base_url(nm_id)

    card = request_json(session, f"{base}/info/ru/card.json", retries=3)
    if not isinstance(card, dict):
        return None

    seller = request_json(session, f"{base}/info/sellers.json", retries=1)
    if not isinstance(seller, dict):
        seller = {}

    price_history = request_json(session, f"{base}/info/price-history.json", retries=1)
    if not isinstance(price_history, list):
        price_history = []

    return StaticBundle(
        base_url=base, card=card, seller=seller, price_history=price_history
    )


def get_price_rub(
    card_v4: dict[str, Any] | None,
    price_history: list[dict[str, Any]],
) -> int | None:
    """Get current price in RUB from v4 card, then fallback to price history."""
    prices: list[int] = []
    if isinstance(card_v4, dict):
        for size in card_v4.get("sizes", []):
            if not isinstance(size, dict):
                continue
            price_data = size.get("price")
            if not isinstance(price_data, dict):
                continue
            product_price = price_data.get("product")
            if isinstance(product_price, (int, float)) and product_price > 0:
                prices.append(int(round(float(product_price) / 100)))
        if prices:
            return min(prices)

    if price_history:
        last = price_history[-1]
        if isinstance(last, dict):
            rub = (last.get("price") or {}).get("RUB")
            if isinstance(rub, (int, float)) and rub > 0:
                return int(round(float(rub) / 100))

    return None


def get_sizes_string(
    card_v4: dict[str, Any] | None,
    card_static: dict[str, Any],
) -> str:
    """Get comma-separated sizes."""
    sizes: list[str] = []

    if isinstance(card_v4, dict):
        for size in card_v4.get("sizes", []):
            if not isinstance(size, dict):
                continue
            value = size.get("origName") or size.get("name")
            if isinstance(value, str) and value.strip():
                sizes.append(value.strip())

    if not sizes:
        values = (card_static.get("sizes_table") or {}).get("values")
        if isinstance(values, list):
            for value in values:
                if not isinstance(value, dict):
                    continue
                tech_size = value.get("tech_size")
                if isinstance(tech_size, str) and tech_size.strip():
                    sizes.append(tech_size.strip())

    deduped = list(dict.fromkeys(sizes))
    return ", ".join(deduped)


def get_stock_quantity(card_v4: dict[str, Any] | None) -> int:
    """Get total stock quantity."""
    if not isinstance(card_v4, dict):
        return 0

    total_quantity = card_v4.get("totalQuantity")
    if isinstance(total_quantity, int):
        return max(total_quantity, 0)

    qty_sum = 0
    for size in card_v4.get("sizes", []):
        if not isinstance(size, dict):
            continue
        for stock in size.get("stocks", []):
            if not isinstance(stock, dict):
                continue
            qty = stock.get("qty")
            if isinstance(qty, int) and qty > 0:
                qty_sum += qty

    return qty_sum


def get_rating_and_reviews(
    card_v4: dict[str, Any] | None,
) -> tuple[float | None, int | None]:
    """Get product rating and review count from v4 card."""
    if not isinstance(card_v4, dict):
        return None, None

    rating: float | None = None
    for key in ("reviewRating", "nmReviewRating", "rating"):
        value = card_v4.get(key)
        if isinstance(value, (int, float)):
            rating = float(value)
            break

    reviews: int | None = None
    for key in ("feedbacks", "nmFeedbacks"):
        value = card_v4.get(key)
        if isinstance(value, int):
            reviews = value
            break

    return rating, reviews


def get_image_links(
    base_url: str,
    card_static: dict[str, Any],
    card_v4: dict[str, Any] | None,
) -> str:
    """Get comma-separated image links."""
    photo_count = 0
    media = card_static.get("media")
    if isinstance(media, dict):
        count = media.get("photo_count")
        if isinstance(count, int):
            photo_count = count

    if not photo_count and isinstance(card_v4, dict):
        pics = card_v4.get("pics")
        if isinstance(pics, int):
            photo_count = pics

    links = [f"{base_url}/images/big/{idx}.webp" for idx in range(1, photo_count + 1)]
    return ", ".join(links)


def build_characteristics(card_static: dict[str, Any]) -> dict[str, Any]:
    """Build structured characteristics payload."""
    result: dict[str, Any] = {}
    for key in (
        "grouped_options",
        "options",
        "compositions",
        "contents",
        "sizes_table",
        "season",
        "kinds",
        "certificate",
    ):
        value = card_static.get(key)
        if value not in (None, "", [], {}):
            result[key] = value
    return result


def extract_country_production(card_static: dict[str, Any]) -> str:
    """Extract 'country of production' from characteristics."""

    def find_country(options: list[dict[str, Any]]) -> str | None:
        for option in options:
            name = str(option.get("name", "")).strip().lower().replace("ё", "е")
            value = str(option.get("value", "")).strip()
            if not value:
                continue
            if "страна" in name and ("производ" in name or "изготов" in name):
                return value
            if "страна производства" in name:
                return value
        return None

    grouped = card_static.get("grouped_options")
    if isinstance(grouped, list):
        for group in grouped:
            options = group.get("options") if isinstance(group, dict) else None
            if isinstance(options, list):
                country = find_country(
                    [opt for opt in options if isinstance(opt, dict)]
                )
                if country:
                    return country

    options = card_static.get("options")
    if isinstance(options, list):
        country = find_country([opt for opt in options if isinstance(opt, dict)])
        if country:
            return country

    return ""


def is_russia(country: str) -> bool:
    """Check if country string indicates Russia."""
    normalized = country.lower().replace("ё", "е")
    return "рос" in normalized or "russia" in normalized


def normalize_text(value: str) -> str:
    """Normalize Cyrillic text for simple substring matching."""
    return value.lower().replace("ё", "е")


def looks_relevant_to_query(card_static: dict[str, Any], query: str) -> bool:
    """Heuristic relevance check for fallback search source."""
    imt_name = str(card_static.get("imt_name", ""))
    subj_name = str(card_static.get("subj_name", ""))
    description = str(card_static.get("description", ""))
    options = json.dumps(card_static.get("options", []), ensure_ascii=False)
    grouped_options = card_static.get("grouped_options", [])
    grouped = json.dumps(grouped_options, ensure_ascii=False)

    text = normalize_text(
        " ".join([imt_name, subj_name, description, options, grouped])
    )
    query_norm = normalize_text(query)

    if "пальто" in query_norm and "шерст" in query_norm:
        return "пальт" in text and "шерст" in text

    words = [word for word in query_norm.split() if len(word) >= 4]
    if not words:
        return True

    stems = []
    for word in words:
        if len(word) > 6:
            stems.append(word[:6])
        else:
            stems.append(word)

    hits = sum(1 for stem in stems if stem in text)
    required_hits = max(1, len(stems) // 2)
    return hits >= required_hits


def build_row(  # pylint: disable=too-many-locals
    nm_id: int,
    static_bundle: StaticBundle,
    card_v4: dict[str, Any] | None,
) -> dict[str, Any]:
    """Build one row for XLSX."""
    card_static = static_bundle.card
    seller_data = static_bundle.seller

    title = card_static.get("imt_name") or (card_v4 or {}).get("name") or ""
    description = card_static.get("description") or ""

    seller_name = (
        seller_data.get("supplierName")
        or (card_v4 or {}).get("supplier")
        or (card_static.get("selling") or {}).get("brand_name")
        or ""
    )

    seller_id = (
        seller_data.get("supplierId")
        or (card_v4 or {}).get("supplierId")
        or (card_static.get("selling") or {}).get("supplier_id")
    )
    seller_url = (
        f"https://www.wildberries.ru/seller/{seller_id}"
        if isinstance(seller_id, int)
        else ""
    )

    price = get_price_rub(card_v4, static_bundle.price_history)
    sizes = get_sizes_string(card_v4, card_static)
    stock = get_stock_quantity(card_v4)
    rating, reviews = get_rating_and_reviews(card_v4)

    characteristics = build_characteristics(card_static)
    characteristics_json = json.dumps(characteristics, ensure_ascii=False)
    country_production = extract_country_production(card_static)

    return {
        "Ссылка на товар": f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
        "Артикул": nm_id,
        "Название": title,
        "Цена": price,
        "Описание": description,
        "Ссылки на изображения через запятую": get_image_links(
            static_bundle.base_url,
            card_static,
            card_v4,
        ),
        "Все характеристики с сохранением структуры": characteristics_json,
        "Название селлера": seller_name,
        "Ссылка на селлера": seller_url,
        "Размеры товара через запятую": sizes,
        "Остатки по товару (число)": stock,
        "Рейтинг": rating,
        "Количество отзывов": reviews,
        "Страна производства": country_production,
    }


def write_xlsx(rows: list[dict[str, Any]], output_file: Path) -> None:
    """Write rows to XLSX file."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "catalog"

    sheet.append(COLUMNS)
    for row in rows:
        sheet.append([row.get(column, "") for column in COLUMNS])

    sheet.freeze_panes = "A2"

    for cell in sheet[1]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    wrap_columns = {
        COLUMNS.index("Описание") + 1,
        COLUMNS.index("Ссылки на изображения через запятую") + 1,
        COLUMNS.index("Все характеристики с сохранением структуры") + 1,
    }

    for col_index in range(1, len(COLUMNS) + 1):
        max_len = len(COLUMNS[col_index - 1])
        for row_index in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row_index, column=col_index).value
            value_len = len(str(value)) if value is not None else 0
            max_len = max(max_len, min(value_len, 120))
            if col_index in wrap_columns:
                sheet.cell(row=row_index, column=col_index).alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

        sheet.column_dimensions[chr(64 + col_index)].width = min(max_len + 2, 70)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)


def filter_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Filter rows by assignment conditions."""
    filtered: list[dict[str, Any]] = []
    for row in rows:
        rating = row.get("Рейтинг")
        price = row.get("Цена")
        country = str(row.get("Страна производства", ""))

        if not isinstance(rating, (int, float)):
            continue
        if not isinstance(price, int):
            continue
        if rating < 4.5:
            continue
        if price > 10000:
            continue
        if not is_russia(country):
            continue

        filtered.append(row)

    return filtered


def parse_arguments() -> argparse.Namespace:
    """Parse command-line arguments."""
    parser = argparse.ArgumentParser(description="WB parser for test assignment")
    parser.add_argument(
        "--query",
        default=DEFAULT_QUERY,
        help="Search query",
    )
    parser.add_argument(
        "--output-all",
        default=DEFAULT_OUTPUT_ALL,
        help="Output XLSX for full catalog",
    )
    parser.add_argument(
        "--output-filtered",
        default=DEFAULT_OUTPUT_FILTERED,
        help="Output XLSX for filtered catalog",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Optional limit by number of products (0 = no limit)",
    )
    parser.add_argument(
        "--sleep",
        type=float,
        default=0.05,
        help="Sleep between products in seconds",
    )
    parser.add_argument(
        "--token",
        default="",
        help="x_wbaas_token (если не передан, будет получен автоматически)",
    )
    parser.add_argument(
        "--chromedriver-path",
        default="",
        help=(
            "Путь к chromedriver. "
            "Если не передан, драйвер будет скачан в ~/.cache/wb_test_parser"
        ),
    )
    parser.add_argument(
        "--token-timeout",
        type=int,
        default=30,
        help="Таймаут ожидания токена в браузере (сек)",
    )
    return parser.parse_args()


def main() -> None:  # pylint: disable=too-many-locals
    """Run parser end-to-end."""
    args = parse_arguments()

    session = requests.Session()
    token = args.token.strip()
    if not token:
        chromedriver_path = ensure_chromedriver(args.chromedriver_path.strip() or None)
        print(f"Получаем токен через Selenium: {chromedriver_path}")
        token = get_wb_token(chromedriver_path, timeout_seconds=args.token_timeout)

    apply_wb_token(session, token)
    short_token = f"{token[:12]}...{token[-8:]}" if len(token) > 24 else token
    print(f"Токен получен и применён: {short_token}")

    ids, source = get_catalog_ids(session, args.query)
    if args.limit > 0:
        ids = ids[: args.limit]

    if not ids:
        raise RuntimeError("Не удалось получить список товаров по запросу")

    print(f"Источник ID: {source}")
    print(f"Найдено ID: {len(ids)}")

    rows: list[dict[str, Any]] = []
    skipped_irrelevant = 0
    for index, nm_id in enumerate(ids, start=1):
        static_bundle = get_static_bundle(session, nm_id)
        if static_bundle is None:
            print(f"[{index}/{len(ids)}] {nm_id}: пропуск (нет static card.json)")
            continue

        if source == "search-goods.wildberries.ru" and not looks_relevant_to_query(
            static_bundle.card, args.query
        ):
            skipped_irrelevant += 1
            continue

        card_v4 = get_card_v4(session, nm_id)
        row = build_row(nm_id, static_bundle, card_v4)
        rows.append(row)

        if index % 25 == 0 or index == len(ids):
            print(f"Прогресс: {index}/{len(ids)}, собрали строк: {len(rows)}")

        if args.sleep > 0:
            time.sleep(args.sleep)

    filtered_rows = filter_rows(rows)

    output_all = Path(args.output_all)
    output_filtered = Path(args.output_filtered)

    write_xlsx(rows, output_all)
    write_xlsx(filtered_rows, output_filtered)

    print(f"Готово: {output_all} (строк: {len(rows)})")
    print(f"Готово: {output_filtered} (строк: {len(filtered_rows)})")
    if skipped_irrelevant:
        print(f"Отфильтровано как нерелевантные запросу: {skipped_irrelevant}")


if __name__ == "__main__":
    main()
